import csv
import re
from io import BytesIO, StringIO

import aiohttp
from aiogram import F, Router
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from bot.config import get_settings
from bot.keyboards.admin import (
    admin_reply_menu,
    admin_products_reply_menu,
    delete_product_confirm_reply_menu,
    deposit_review_reply_menu,
    export_reply_menu,
    member_actions_reply_menu,
    member_orders_reply_menu,
    members_reply_menu,
    paged_reply_menu,
    product_admin_actions_reply_menu,
    refund_confirm_reply_menu,
    replacement_review_reply_menu,
)
from bot.services.coupons import create_coupon
from bot.services.audit import log_admin_action
from bot.services.deposits import all_deposits, deposited_today, pending_deposits, review_deposit
from bot.services.products import add_stock, add_stock_batch, create_product, delete_product, list_all_products, search_products, toggle_product, unsold_stock_count, unsold_stock_items, update_product
from bot.services.auto_stock import (
    DEFAULT_REFILL_THRESHOLD,
    DEFAULT_TARGET_STOCK,
    get_auto_stock_source,
    refill_source,
    reset_auto_stock_progress,
    stop_auto_stock_source,
    upsert_auto_stock_source,
)
from bot.services.stats import admin_stats
from bot.services.orders import all_orders, get_order, order_count, refund_order, sales_report, total_spent
from bot.services.replacements import pending_replacements, review_replacement
from bot.services.users import adjust_user_balance, find_user, list_all_users, list_recent_users, set_user_banned, set_user_note, set_user_restricted
from bot.utils.formatting import money

router = Router()

SUPPORTED_STOCK_EXTENSIONS = (".xlsx", ".csv", ".txt")
MAX_TELEGRAM_DOWNLOAD_BYTES = 20 * 1024 * 1024


class ProductForm(StatesGroup):
    details = State()


class ProductEditForm(StatesGroup):
    details = State()


class StockForm(StatesGroup):
    product_id = State()
    payload = State()


class StockUrlForm(StatesGroup):
    url = State()


class AutoStockForm(StatesGroup):
    details = State()


class CouponAdminForm(StatesGroup):
    details = State()


class MemberLookupForm(StatesGroup):
    query = State()


class MemberBalanceForm(StatesGroup):
    amount = State()


class MemberNoteForm(StatesGroup):
    note = State()


class BroadcastForm(StatesGroup):
    message = State()


class AdminSearchForm(StatesGroup):
    query = State()


class DepositRejectReasonForm(StatesGroup):
    reason = State()


class RefundReasonForm(StatesGroup):
    reason = State()


def _is_admin_document_message(message: Message) -> bool:
    settings = get_settings()
    admin_ids = settings.admin_ids | settings.support_admin_ids | settings.stock_manager_ids
    return bool(message.document and message.from_user and message.from_user.id in admin_ids)


@router.message(StateFilter("*"), _is_admin_document_message)
async def admin_document_catcher(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await _process_stock_document(message, state, session)


def is_admin(user_id: int) -> bool:
    settings = get_settings()
    return user_id in settings.admin_ids or user_id in settings.support_admin_ids or user_id in settings.stock_manager_ids


def is_owner_admin(user_id: int) -> bool:
    return user_id in get_settings().admin_ids


def _button_text(text: str | None) -> str:
    if not text:
        return ""
    return re.sub(r"^[^A-Za-z0-9#]+", "", text.strip()).strip()


def _starts_with_any(text: str, prefixes: tuple[str, ...]) -> bool:
    normalized = _button_text(text)
    upper_normalized = normalized.upper()
    return any(normalized.startswith(prefix) or upper_normalized.startswith(prefix.upper()) for prefix in prefixes)


def _is_product_selection(text: str) -> bool:
    return _starts_with_any(text, ("Product #",))


def _is_product_id_selection(text: str | None) -> bool:
    normalized = _button_text(text)
    return normalized.isdigit()


def _is_add_stock_action(text: str) -> bool:
    return _starts_with_any(text, ("Add Stock #",))


def _is_export_stock_action(text: str) -> bool:
    return _starts_with_any(text, ("Export Stock #",))


def _is_import_stock_url_action(text: str) -> bool:
    return _starts_with_any(text, ("Import Stock URL #",))


def _is_auto_refill_action(text: str) -> bool:
    return _starts_with_any(text, ("Auto Refill #",))


def _is_auto_status_action(text: str) -> bool:
    return _starts_with_any(text, ("Auto Status #",))


def _is_stop_auto_action(text: str) -> bool:
    return _starts_with_any(text, ("Stop Auto #",))


def _is_reset_auto_action(text: str) -> bool:
    return _starts_with_any(text, ("Reset Auto #",))


def _is_exact_button(text: str, label: str) -> bool:
    return _button_text(text).casefold() == label.casefold()


def _is_edit_product_action(text: str) -> bool:
    return _starts_with_any(text, ("Edit Product #",))


def _is_product_action(text: str) -> bool:
    return _starts_with_any(
        text,
        (
            "Enable Product #",
            "Disable Product #",
            "Delete Product #",
            "Cancel Product #",
        ),
    )


def _is_delete_confirm(text: str) -> bool:
    return _starts_with_any(text, ("Confirm Delete Product #",))


def _is_deposit_review(text: str) -> bool:
    return _starts_with_any(text, ("Approve Deposit #", "Reject Deposit #"))


def _is_deposit_approve(text: str) -> bool:
    return _starts_with_any(text, ("Approve Deposit #",))


def _is_replacement_review(text: str) -> bool:
    return _starts_with_any(text, ("Approve Replace #", "Reject Replace #"))


def _is_replacement_approve(text: str) -> bool:
    return _starts_with_any(text, ("Approve Replace #",))


def _is_member_action(text: str) -> bool:
    return _starts_with_any(
        text,
        (
            "Add Balance #",
            "Remove Balance #",
            "Check Orders #",
            "Check Balance #",
            "Note Member #",
            "Ban Member #",
            "Unban Member #",
            "Restrict Member #",
            "Unrestrict Member #",
        ),
    )


def _is_refund_order(text: str) -> bool:
    return _starts_with_any(text, ("Refund Order #",))


def _is_refund_confirm(text: str) -> bool:
    return _starts_with_any(text, ("Confirm Refund Order #",))


def _is_member_selection(text: str) -> bool:
    return _starts_with_any(text, ("Member #",))


def _page_from_text(text: str, prefix: str) -> int | None:
    normalized = _button_text(text)
    if not (normalized.startswith(f"{prefix} Page ") or normalized.startswith(f"{prefix} PAGE ")):
        return None
    match = re.search(r"Page\s+(\d+)", normalized, flags=re.IGNORECASE)
    return int(match.group(1)) if match else None


def _page_slice(items: list[object], page: int, per_page: int = 10) -> list[object]:
    start = max(page - 1, 0) * per_page
    return items[start : start + per_page]


def _member_status(user: object) -> str:
    flags = []
    if getattr(user, "is_banned", False):
        flags.append("Banned")
    if getattr(user, "is_restricted", False):
        flags.append("Restricted")
    return ", ".join(flags) if flags else "Normal"


def _member_text(user: object) -> str:
    username = f"@{user.username}" if user.username else "not_available"
    return (
        "Member Details\n\n"
        f"Database ID: #{user.id}\n"
        f"Name: {user.first_name or 'Unknown'}\n"
        f"Username: {username}\n"
        f"Telegram ID: <code>{user.telegram_id}</code>\n"
        f"Balance: {money(user.balance)}\n"
        f"Status: {_member_status(user)}\n"
        f"Joined: {user.created_at:%Y-%m-%d %H:%M}\n"
        f"Admin Note: {user.admin_note or 'None'}"
    )


async def _member_activity_text(session: AsyncSession, user: object) -> str:
    orders = await order_count(session, user.id)
    spent = await total_spent(session, user.id)
    today_deposit = await deposited_today(session, user.id)
    return (
        f"{_member_text(user)}\n"
        f"Total Orders: {orders}\n"
        f"Total Spent: {money(spent)}\n"
        f"Deposited Today: {money(today_deposit)}"
    )


def _id_from_hash_button(text: str, prefix: str) -> int | None:
    normalized = _button_text(text)
    if not normalized.startswith(prefix):
        return None
    match = re.search(r"#\s*(\d+)", normalized)
    if not match:
        return None
    return int(match.group(1))


async def _admin_dashboard_text(session: AsyncSession) -> str:
    stats = await admin_stats(session)
    today = await sales_report(session, days=1)
    return (
        "Admin Dashboard\n\n"
        f"Today Orders: {today['orders']}\n"
        f"Today Revenue: {money(today['revenue'])}\n"
        f"Pending Deposits: {stats['pending_deposits']}\n"
        f"Total Users: {stats['users']}\n"
        f"Products: {stats['products']}\n"
        f"Unsold Stock: {stats['stock']}"
    )


def _method_label(method: str) -> str:
    labels = {
        "binance": "Binance",
        "usdt_trc20": "USDT TRC20",
        "usdt_bep20": "USDT BEP20",
        "bkash": "bKash",
        "nagad": "Nagad",
        "rocket": "Rocket",
    }
    return labels.get(method, method.upper())


async def _deposit_details_text(session: AsyncSession, deposit: object, queue_size: int | None = None) -> str:
    from bot.database.models import User

    user = await session.get(User, deposit.user_id)
    queue_line = f"\nQueue Size: {queue_size}" if queue_size is not None else ""
    username = f"@{user.username}" if user and user.username else "not_available"
    name = user.first_name if user and user.first_name else "Unknown"
    telegram_id = user.telegram_id if user else "Unknown"
    return (
        "Pending Deposit Request\n\n"
        f"Request ID: #{deposit.id}\n"
        f"Amount: {money(deposit.amount)}\n"
        f"Method: {_method_label(deposit.method)}\n"
        f"Transaction ID: <code>{deposit.transaction_id}</code>\n"
        f"Screenshot: {'Attached' if getattr(deposit, 'proof_file_id', None) else 'Missing'}\n"
        f"OCR Status: {getattr(deposit, 'ocr_status', None) or 'Not checked'}\n"
        f"Status: {deposit.status.value}\n"
        f"Created: {deposit.created_at:%Y-%m-%d %H:%M}"
        f"{queue_line}\n\n"
        "Customer Information\n"
        f"Name: {name}\n"
        f"Username: {username}\n"
        f"Telegram ID: <code>{telegram_id}</code>\n\n"
        "OCR Assistant\n"
        f"{getattr(deposit, 'ocr_details', None) or 'No OCR details available.'}\n\n"
        "Verify the payment manually, then approve or reject."
    )


async def _send_next_deposit_after_review(message: Message, session: AsyncSession, deposit: object) -> None:
    await message.answer(
        f"Deposit Reviewed\n\nRequest ID: #{deposit.id}\nStatus: {deposit.status.value}\nAmount: {money(deposit.amount)}"
    )
    rows = await pending_deposits(session)
    if not rows:
        await message.answer("Deposits\n\nNo pending deposit requests.", reply_markup=admin_reply_menu())
        return
    first = rows[0]
    text = await _deposit_details_text(session, first, len(rows))
    if getattr(first, "proof_file_id", None):
        await message.answer_photo(first.proof_file_id, caption=f"Payment screenshot for deposit #{first.id}")
    await message.answer(text, reply_markup=deposit_review_reply_menu(first.id))


async def _replacement_details_text(session: AsyncSession, request: object, queue_size: int | None = None) -> str:
    from bot.database.models import User

    user = await session.get(User, request.user_id)
    queue_line = f"Queue: {queue_size} pending\n" if queue_size is not None else ""
    return (
        "Replacement Review\n\n"
        f"{queue_line}"
        f"Request ID: #{request.id}\n"
        f"User: {user.first_name if user else 'Unknown'}\n"
        f"Username: @{user.username if user and user.username else 'not_available'}\n"
        f"Telegram ID: <code>{user.telegram_id if user else 'unknown'}</code>\n"
        f"Order ID: #{request.order_id or 'Not given'}\n"
        f"Quantity: {request.quantity}\n"
        f"Status: {request.status.value}\n"
        f"Submitted: {request.created_at:%Y-%m-%d %H:%M}\n\n"
        f"Message:\n{request.message}\n\n"
        f"Proof: {'File/photo attached' if request.proof_file_id else 'Text proof'}"
    )


async def _send_next_replacement_after_review(message: Message, session: AsyncSession, request: object) -> None:
    await message.answer(
        f"Replacement Reviewed\n\nRequest ID: #{request.id}\nStatus: {request.status.value}\nRefund: {money(request.refund_amount)}"
    )
    rows = await pending_replacements(session)
    if not rows:
        await message.answer("Replacements\n\nNo pending replacement requests.", reply_markup=admin_reply_menu())
        return
    first = rows[0]
    text = await _replacement_details_text(session, first, len(rows))
    if getattr(first, "proof_file_id", None) and getattr(first, "proof_file_name", None):
        await message.answer_document(first.proof_file_id, caption=f"Replacement proof for request #{first.id}")
    elif getattr(first, "proof_file_id", None):
        await message.answer_photo(first.proof_file_id, caption=f"Replacement proof for request #{first.id}")
    elif getattr(first, "proof_text", None):
        await message.answer(f"Proof text #{first.id}:\n\n{first.proof_text[:3500]}")
    await message.answer(text, reply_markup=replacement_review_reply_menu(first.id))


def _looks_like_header(values: list[str]) -> bool:
    header_words = {"email", "mail", "username", "user", "password", "pass", "account", "accounts"}
    lowered = {value.strip().lower() for value in values if value.strip()}
    return bool(lowered & header_words)


def _looks_like_stock_line(value: str) -> bool:
    clean = value.strip()
    return "@" in clean and ("|" in clean or ":" in clean)


def _stock_line_from_cells(cells: list[object]) -> str | None:
    values = [str(cell).strip() for cell in cells if cell is not None and str(cell).strip()]
    if not values or _looks_like_header(values):
        return None
    if values[0].isdigit() and len(values) >= 4 and _looks_like_stock_line(values[-1]):
        return values[-1]
    if len(values) >= 4 and values[0].lower() in {"no", "#"} and _looks_like_stock_line(values[-1]):
        return values[-1]
    if len(values) >= 2:
        if len(values) > 2 and any("|" in value for value in values[1:]):
            return "|".join(values)
        return f"{values[0]}|{values[1]}"
    return values[0]


def parse_stock_file(file_name: str, data: bytes) -> list[str]:
    lowered = file_name.lower()
    if lowered.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        worksheet = workbook.active
        lines = []
        for row in worksheet.iter_rows(values_only=True):
            line = _stock_line_from_cells(list(row))
            if line:
                lines.append(line)
        workbook.close()
        return lines

    if lowered.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="ignore")
        rows = csv.reader(StringIO(text))
        return [line for row in rows if (line := _stock_line_from_cells(row))]

    if lowered.endswith(".txt"):
        text = data.decode("utf-8-sig", errors="ignore")
        return [line for raw_line in text.splitlines() if (line := _stock_line_from_text(raw_line))]

    raise ValueError("Unsupported file type.")


def _normalize_stock_url(url: str) -> str:
    cleaned = url.strip()
    if "dropbox.com" in cleaned:
        cleaned = cleaned.replace("?dl=0", "?dl=1")
        if "?dl=" not in cleaned:
            cleaned += ("&" if "?" in cleaned else "?") + "dl=1"
    return cleaned


def _stock_line_from_text(raw_line: str) -> str | None:
    line = raw_line.strip()
    if not line:
        return None
    if "\t" in line:
        row = line.split("\t")
        return _stock_line_from_cells(row)
    if "|" in line:
        return line
    try:
        row = next(csv.reader([line]))
    except Exception:
        return line
    return _stock_line_from_cells(row) or line


async def import_stock_from_url(
    session: AsyncSession,
    product_id: int,
    url: str,
    batch_size: int = 5000,
) -> int:
    total = 0
    pending: list[str] = []
    remainder = ""
    timeout = aiohttp.ClientTimeout(total=None, sock_connect=30, sock_read=300)
    async with aiohttp.ClientSession(timeout=timeout) as http_session:
        async with http_session.get(_normalize_stock_url(url), allow_redirects=True) as response:
            if response.status >= 400:
                raise ValueError(f"Download failed with HTTP {response.status}.")
            async for chunk in response.content.iter_chunked(1024 * 256):
                text = remainder + chunk.decode("utf-8-sig", errors="ignore")
                lines = text.splitlines()
                if text and not text.endswith(("\n", "\r")):
                    remainder = lines.pop() if lines else text
                else:
                    remainder = ""
                for raw_line in lines:
                    line = _stock_line_from_text(raw_line)
                    if line:
                        pending.append(line)
                    if len(pending) >= batch_size:
                        total += await add_stock_batch(session, product_id, pending)
                        pending = []
            if remainder:
                line = _stock_line_from_text(remainder)
                if line:
                    pending.append(line)
            if pending:
                total += await add_stock_batch(session, product_id, pending)
    return total


def _xlsx_file(filename: str, rows: list[list[object]], sheet_name: str = "Export") -> BufferedInputFile:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return BufferedInputFile(output.read(), filename=filename)


@router.message(Command("admin"))
async def admin_command(message: Message, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await message.answer(
        await _admin_dashboard_text(session),
        reply_markup=admin_reply_menu(),
    )


@router.callback_query(F.data == "admin")
async def admin_callback(callback: CallbackQuery, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    await callback.message.edit_text("Admin Panel")
    await callback.message.answer(
        await _admin_dashboard_text(session),
        reply_markup=admin_reply_menu(),
    )
    await callback.answer()


@router.message(StateFilter("*"), F.text.in_({"Admin Panel", "ADMIN PANEL", "⚙️ Admin Panel"}))
async def admin_panel_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await message.answer(
        await _admin_dashboard_text(session),
        reply_markup=admin_reply_menu(),
    )


@router.callback_query(F.data == "admin_stats")
async def stats(callback: CallbackQuery, session: AsyncSession) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    data = await admin_stats(session)
    await callback.message.edit_text(
        "Store Statistics\n\n"
        f"Users: {data['users']}\n"
        f"Products: {data['products']}\n"
        f"Unsold stock: {data['stock']}\n"
        f"Orders: {data['orders']}\n"
        f"Revenue: {money(data['revenue'])}\n"
        f"Pending deposits: {data['pending_deposits']}",
    )
    await callback.message.answer("Admin Panel", reply_markup=admin_reply_menu())
    await callback.answer()


@router.message(StateFilter("*"), F.text.in_({"Stats", "STATS", "📊 Stats"}))
async def stats_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    data = await admin_stats(session)
    await message.answer(
        "Store Statistics\n\n"
        f"Users: {data['users']}\n"
        f"Products: {data['products']}\n"
        f"Unsold stock: {data['stock']}\n"
        f"Orders: {data['orders']}\n"
        f"Revenue: {money(data['revenue'])}\n"
        f"Pending deposits: {data['pending_deposits']}",
        reply_markup=admin_reply_menu(),
    )


@router.message(StateFilter("*"), F.text.in_({"Reports", "REPORTS", "📈 Reports"}))
async def reports_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    today = await sales_report(session, days=1)
    week = await sales_report(session, days=7)
    month = await sales_report(session, days=30)
    await message.answer(
        "Sales Reports\n\n"
        f"Today: {today['orders']} orders | Revenue {money(today['revenue'])} | Refunded {money(today['refunded'])}\n"
        f"7 Days: {week['orders']} orders | Revenue {money(week['revenue'])} | Refunded {money(week['refunded'])}\n"
        f"30 Days: {month['orders']} orders | Revenue {money(month['revenue'])} | Refunded {money(month['refunded'])}",
        reply_markup=admin_reply_menu(),
    )


@router.message(StateFilter("*"), F.text.in_({"Broadcast", "BROADCAST", "📣 Broadcast"}))
async def broadcast_start_text(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await state.set_state(BroadcastForm.message)
    await message.answer("Broadcast\n\nSend the message you want to send to all members.")


@router.message(BroadcastForm.message)
async def broadcast_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    users = await list_all_users(session)
    sent = 0
    failed = 0
    for user in users:
        try:
            await message.bot.send_message(user.telegram_id, message.text)
            sent += 1
        except Exception:
            failed += 1
    await state.clear()
    await log_admin_action(session, message.from_user.id, "broadcast", details=f"sent={sent}, failed={failed}")
    await message.answer(f"Broadcast completed.\n\nSent: {sent}\nFailed: {failed}", reply_markup=admin_reply_menu())


@router.message(StateFilter("*"), F.text.in_({"Admin Search", "🔎 Admin Search"}))
async def admin_search_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await state.set_state(AdminSearchForm.query)
    await message.answer(
        "Admin Search\n\n"
        "Send one of these:\n"
        "user 7562995992\n"
        "product gmail\n"
        "order 12"
    )


@router.message(AdminSearchForm.query)
async def admin_search_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    text = message.text.strip()
    if text.lower().startswith("user "):
        user = await find_user(session, text.split(" ", 1)[1])
        if not user:
            await message.answer("User not found.", reply_markup=admin_reply_menu())
            return
        await message.answer(
            await _member_activity_text(session, user),
            reply_markup=member_actions_reply_menu(user.id, user.is_banned, user.is_restricted),
        )
        return
    if text.lower().startswith("product "):
        rows = await search_products(session, text.split(" ", 1)[1])
        if not rows:
            await message.answer("Product not found.", reply_markup=admin_reply_menu())
            return
        response = "Product Search\n\n" + "\n".join(
            f"#{product.id} {product.name} - {money(product.price)} - stock {stock} - {'active' if product.is_active else 'disabled'}"
            for product, stock in rows
        )
        await message.answer(response, reply_markup=admin_products_reply_menu(rows))
        return
    if text.lower().startswith("order "):
        raw_id = text.split(" ", 1)[1].strip()
        if not raw_id.isdigit():
            await message.answer("Send order ID like: order 12", reply_markup=admin_reply_menu())
            return
        order = await get_order(session, int(raw_id))
        if not order:
            await message.answer("Order not found.", reply_markup=admin_reply_menu())
            return
        await message.answer(
            "Order Details\n\n"
            f"Order ID: #{order.id}\n"
            f"User ID: #{order.user_id}\n"
            f"Product ID: #{order.product_id}\n"
            f"Amount: {money(order.amount)}\n"
            f"Status: {order.status.value}\n"
            f"Created: {order.created_at:%Y-%m-%d %H:%M}",
            reply_markup=refund_confirm_reply_menu(order.id) if order.status.value != "refunded" else admin_reply_menu(),
        )
        return
    await message.answer("Unknown search. Use: user ..., product ..., or order ...", reply_markup=admin_reply_menu())


@router.message(StateFilter("*"), F.text.func(lambda text: _is_exact_button(text, "Export Data")))
async def export_menu_text(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await message.answer("Export Data\n\nSelect what you want to export.", reply_markup=export_reply_menu())


@router.message(StateFilter("*"), F.text.func(lambda text: _is_exact_button(text, "Export Users")))
async def export_users_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await message.answer("Preparing users export...")
    users = await list_all_users(session)
    rows = [["DB ID", "Telegram ID", "Name", "Username", "Balance", "Status", "Note", "Joined"]]
    for user in users:
        rows.append([user.id, user.telegram_id, user.first_name or "", user.username or "", float(user.balance), _member_status(user), user.admin_note or "", str(user.created_at)])
    await message.answer_document(_xlsx_file("users_export.xlsx", rows, "Users"), caption="Users export ready.", reply_markup=admin_reply_menu())


@router.message(StateFilter("*"), F.text.func(lambda text: _is_exact_button(text, "Export Orders")))
async def export_orders_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await message.answer("Preparing orders export...")
    orders = await all_orders(session)
    rows = [["Order ID", "User ID", "Product ID", "Amount", "Status", "Created"]]
    for order in orders:
        rows.append([order.id, order.user_id, order.product_id, float(order.amount), order.status.value, str(order.created_at)])
    await message.answer_document(_xlsx_file("orders_export.xlsx", rows, "Orders"), caption="Orders export ready.", reply_markup=admin_reply_menu())


@router.message(StateFilter("*"), F.text.func(lambda text: _is_exact_button(text, "Export Deposits")))
async def export_deposits_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await message.answer("Preparing deposits export...")
    deposits = await all_deposits(session)
    rows = [["Deposit ID", "User ID", "Amount", "Method", "TXID", "Status", "OCR", "Created", "Reviewed"]]
    for deposit in deposits:
        rows.append([deposit.id, deposit.user_id, float(deposit.amount), deposit.method, deposit.transaction_id or "", deposit.status.value, deposit.ocr_status or "", str(deposit.created_at), str(deposit.reviewed_at or "")])
    await message.answer_document(_xlsx_file("deposits_export.xlsx", rows, "Deposits"), caption="Deposits export ready.", reply_markup=admin_reply_menu())


@router.message(StateFilter("*"), F.text.func(lambda text: _is_exact_button(text, "Export Products")))
async def export_products_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await message.answer("Preparing products export...")
    products = await list_all_products(session)
    rows = [["Product ID", "Name", "Price", "Unsold Stock", "Active", "Description"]]
    for product, stock in products:
        rows.append([product.id, product.name, float(product.price), stock, product.is_active, product.description or ""])
    await message.answer_document(_xlsx_file("products_export.xlsx", rows, "Products"), caption="Products export ready.", reply_markup=admin_reply_menu())


@router.message(StateFilter("*"), F.text.in_({"Members", "MEMBERS", "👥 Members", "Search Member", "🔎 Search Member"}))
async def members_start_text(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await state.set_state(MemberLookupForm.query)
    await message.answer(
        "Member Management\n\n"
        "Send member Telegram ID or username.\n\n"
        "Example:\n"
        "7562995992\n"
        "@username"
    )


@router.message(StateFilter("*"), F.text.in_({"All Members", "📋 All Members"}))
async def all_members_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    users = _page_slice(await list_all_users(session), page=1, per_page=10)
    if not users:
        await message.answer("Members\n\nNo members found.", reply_markup=admin_reply_menu())
        return
    lines = []
    for user in users:
        username = f"@{user.username}" if user.username else "no_username"
        lines.append(
            f"#{user.id} | {user.first_name or 'Unknown'} | {username} | TG: {user.telegram_id} | {money(user.balance)} | {_member_status(user)}"
        )
    await message.answer(
        "All Members\n\n"
        + "\n".join(lines)
        + "\n\nSelect a member from the keyboard below.",
        reply_markup=members_reply_menu(users),
    )


@router.message(StateFilter("*"), F.text.func(lambda text: _page_from_text(text, "Members") is not None))
async def members_page_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    page = _page_from_text(message.text, "Members") or 1
    users = _page_slice(await list_all_users(session), page=page, per_page=10)
    if not users:
        await message.answer("No members on this page.", reply_markup=paged_reply_menu("Members", page))
        return
    lines = []
    for user in users:
        username = f"@{user.username}" if user.username else "no_username"
        lines.append(f"#{user.id} | {user.first_name or 'Unknown'} | {username} | TG: {user.telegram_id} | {money(user.balance)} | {_member_status(user)}")
    await message.answer(
        f"All Members - Page {page}\n\n" + "\n".join(lines),
        reply_markup=members_reply_menu(users),
    )
    await message.answer("Navigate member pages.", reply_markup=paged_reply_menu("Members", page))


@router.message(StateFilter("*"), F.text.func(_is_member_selection))
async def member_select_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    user_id = _id_from_hash_button(message.text, "Member #")
    if not user_id:
        await message.answer("Invalid member selection.", reply_markup=admin_reply_menu())
        return
    from bot.database.models import User

    user = await session.get(User, user_id)
    if not user:
        await message.answer("Member not found.", reply_markup=admin_reply_menu())
        return
    await message.answer(
        await _member_activity_text(session, user),
        reply_markup=member_actions_reply_menu(user.id, user.is_banned, user.is_restricted),
    )


@router.message(MemberLookupForm.query)
async def member_lookup_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    user = await find_user(session, message.text)
    await state.clear()
    if not user:
        await message.answer("Member not found.", reply_markup=admin_reply_menu())
        return
    await message.answer(
        await _member_activity_text(session, user),
        reply_markup=member_actions_reply_menu(user.id, user.is_banned, user.is_restricted),
    )


@router.message(StateFilter("*"), F.text.func(_is_member_action))
async def member_action_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return

    action_text = _button_text(message.text)
    user_id = None
    for prefix in (
        "Add Balance #",
        "Remove Balance #",
        "Check Orders #",
        "Check Balance #",
        "Note Member #",
        "Ban Member #",
        "Unban Member #",
        "Restrict Member #",
        "Unrestrict Member #",
    ):
        user_id = _id_from_hash_button(message.text, prefix)
        if user_id:
            break
    if not user_id:
        await message.answer("Invalid member action.", reply_markup=admin_reply_menu())
        return

    from bot.database.models import User

    user = await session.get(User, user_id)
    if not user:
        await message.answer("Member not found.", reply_markup=admin_reply_menu())
        return
    if user.telegram_id in get_settings().admin_ids and (
        action_text.startswith("Ban Member #")
        or action_text.startswith("Restrict Member #")
    ):
        await message.answer("Admin accounts cannot be banned or restricted.", reply_markup=admin_reply_menu())
        return

    if action_text.startswith("Add Balance #") or action_text.startswith("Remove Balance #"):
        action = "add" if action_text.startswith("Add Balance #") else "remove"
        await state.update_data(member_id=user.id, balance_action=action)
        await state.set_state(MemberBalanceForm.amount)
        await message.answer(
            f"{'Add' if action == 'add' else 'Remove'} Balance\n\n"
            f"Member: {user.first_name or user.telegram_id}\n"
            f"Current Balance: {money(user.balance)}\n\n"
            "Send amount."
        )
        return

    if action_text.startswith("Check Balance #"):
        await message.answer(
            await _member_activity_text(session, user),
            reply_markup=member_actions_reply_menu(user.id, user.is_banned, user.is_restricted),
        )
        return

    if action_text.startswith("Note Member #"):
        await state.update_data(member_id=user.id)
        await state.set_state(MemberNoteForm.note)
        await message.answer(
            "Admin Note\n\n"
            f"Member: {user.first_name or user.telegram_id}\n"
            f"Current Note: {user.admin_note or 'None'}\n\n"
            "Send new note. Send '-' to clear note."
        )
        return

    if action_text.startswith("Check Orders #"):
        from bot.services.orders import recent_orders

        orders = await recent_orders(session, user.id, limit=10)
        if not orders:
            text = "Member Orders\n\nNo orders found."
        else:
            text = "Member Orders\n\n" + "\n".join(
                f"#{order.id} - {money(order.amount)} - {order.status.value} - {order.created_at:%Y-%m-%d %H:%M}"
                for order in orders
            )
        await message.answer(text, reply_markup=member_orders_reply_menu(user.id, orders))
        return

    if action_text.startswith("Ban Member #") or action_text.startswith("Unban Member #"):
        banned = action_text.startswith("Ban Member #")
        user = await set_user_banned(session, user.id, banned)
        await log_admin_action(session, message.from_user.id, "ban_member" if banned else "unban_member", "user", user.id)
        await message.answer(
            f"Member {'Banned' if banned else 'Unbanned'}\n\n{await _member_activity_text(session, user)}",
            reply_markup=member_actions_reply_menu(user.id, user.is_banned, user.is_restricted),
        )
        return

    if action_text.startswith("Restrict Member #") or action_text.startswith("Unrestrict Member #"):
        restricted = action_text.startswith("Restrict Member #")
        user = await set_user_restricted(session, user.id, restricted)
        await log_admin_action(session, message.from_user.id, "restrict_member" if restricted else "unrestrict_member", "user", user.id)
        await message.answer(
            f"Member {'Restricted' if restricted else 'Unrestricted'}\n\n{await _member_activity_text(session, user)}",
            reply_markup=member_actions_reply_menu(user.id, user.is_banned, user.is_restricted),
        )


@router.message(MemberBalanceForm.amount)
async def member_balance_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    try:
        amount = float(message.text)
    except (TypeError, ValueError):
        await message.answer("Please send a valid numeric amount.")
        return
    if amount <= 0:
        await message.answer("Amount must be greater than 0.")
        return

    data = await state.get_data()
    signed_amount = amount if data["balance_action"] == "add" else -amount
    user = await adjust_user_balance(session, int(data["member_id"]), signed_amount)
    await state.clear()
    if not user:
        await message.answer("Member not found.", reply_markup=admin_reply_menu())
        return
    await message.answer(
        f"Balance Updated\n\n"
        f"Member: {user.first_name or user.telegram_id}\n"
        f"New Balance: {money(user.balance)}",
        reply_markup=member_actions_reply_menu(user.id, user.is_banned, user.is_restricted),
    )
    await log_admin_action(
        session,
        message.from_user.id,
        "add_balance" if signed_amount > 0 else "remove_balance",
        "user",
        user.id,
        details=f"amount={amount}",
    )


@router.message(MemberNoteForm.note)
async def member_note_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    note = "" if message.text.strip() == "-" else message.text
    user = await set_user_note(session, int(data["member_id"]), note)
    await state.clear()
    if not user:
        await message.answer("Member not found.", reply_markup=admin_reply_menu())
        return
    await log_admin_action(session, message.from_user.id, "set_member_note", "user", user.id)
    await message.answer(
        f"Note Updated\n\n{await _member_activity_text(session, user)}",
        reply_markup=member_actions_reply_menu(user.id, user.is_banned, user.is_restricted),
    )


@router.message(StateFilter("*"), F.text.func(_is_refund_order))
async def refund_order_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    order_id = _id_from_hash_button(message.text, "Refund Order #")
    if not order_id:
        await message.answer("Invalid order action.", reply_markup=admin_reply_menu())
        return
    await message.answer(
        f"Confirm refund for order #{order_id}?\n\nThis will mark the order as refunded and return the amount to the member balance.",
        reply_markup=refund_confirm_reply_menu(order_id),
    )


@router.message(StateFilter("*"), F.text.func(_is_refund_confirm))
async def refund_order_confirm_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    order_id = _id_from_hash_button(message.text, "Confirm Refund Order #")
    if not order_id:
        await message.answer("Invalid order action.", reply_markup=admin_reply_menu())
        return
    await state.update_data(order_id=order_id)
    await state.set_state(RefundReasonForm.reason)
    await message.answer(
        f"Refund Reason\n\nSend reason for order #{order_id} refund.\n\nExample: duplicate order / stock issue / manual adjustment"
    )


@router.message(RefundReasonForm.reason)
async def refund_reason_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    reason = message.text.strip()
    if not reason:
        await message.answer("Please send a refund reason.")
        return
    order_id = int(data["order_id"])
    ok, text, order = await refund_order(session, order_id)
    await state.clear()
    if order:
        await log_admin_action(session, message.from_user.id, "refund_order", "order", order.id, details=f"{text}; reason={reason}")
        from bot.database.models import User

        user = await session.get(User, order.user_id)
        if user:
            try:
                await message.bot.send_message(
                    user.telegram_id,
                    "Order Refund\n\n"
                    f"Order ID: #{order.id}\n"
                    f"Refunded Amount: {money(order.amount)}\n"
                    f"Reason: {reason}\n"
                    "Your balance has been updated.",
                )
            except Exception:
                pass
    await message.answer(text, reply_markup=admin_reply_menu())


@router.callback_query(F.data == "admin_products")
async def admin_products(callback: CallbackQuery, session: AsyncSession) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    rows = await list_all_products(session)
    if not rows:
        await callback.message.edit_text("No products created.")
        await callback.message.answer("Admin Panel", reply_markup=admin_reply_menu())
    else:
        text = "Product List\n\n" + "\n".join(
            f"#{product.id} {product.name} - {money(product.price)} - stock {stock} - {'active' if product.is_active else 'disabled'}"
            for product, stock in rows
        )
        await callback.message.edit_text(text)
        await callback.message.answer("Select a product from the keyboard.", reply_markup=admin_products_reply_menu(rows))
    await callback.answer()


@router.message(StateFilter("*"), F.text.in_({"Products", "PRODUCTS", "📦 Products"}))
async def admin_products_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    rows = _page_slice(await list_all_products(session), page=1, per_page=10)
    if not rows:
        await message.answer("No products created.", reply_markup=admin_reply_menu())
    else:
        text = "Product List\n\n" + "\n".join(
            f"#{product.id} {product.name} - {money(product.price)} - stock {stock} - {'active' if product.is_active else 'disabled'}"
            for product, stock in rows
        )
        await message.answer(text, reply_markup=admin_products_reply_menu(rows))


@router.message(StateFilter("*"), F.text.func(lambda text: _page_from_text(text, "Products") is not None))
async def products_page_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    page = _page_from_text(message.text, "Products") or 1
    rows = _page_slice(await list_all_products(session), page=page, per_page=10)
    if not rows:
        await message.answer("No products on this page.", reply_markup=paged_reply_menu("Products", page))
        return
    text = f"Product List - Page {page}\n\n" + "\n".join(
        f"#{product.id} {product.name} - {money(product.price)} - stock {stock} - {'active' if product.is_active else 'disabled'}"
        for product, stock in rows
    )
    await message.answer(text, reply_markup=admin_products_reply_menu(rows))


@router.message(StateFilter(None), F.text.func(_is_product_id_selection))
async def admin_product_id_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = int(_button_text(message.text))
    rows = await list_all_products(session)
    product_row = next((row for row in rows if row[0].id == product_id), None)
    if not product_row:
        await message.answer("Product not found.", reply_markup=admin_reply_menu())
        return
    product, stock_count = product_row
    await message.answer(
        f"Product Details\n\n"
        f"Product ID: #{product.id}\n"
        f"Name: {product.name}\n"
        f"Price: {money(product.price)}\n"
        f"Available Stock: {stock_count}\n"
        f"Status: {'Active' if product.is_active else 'Disabled'}\n\n"
        f"Description: {product.description or 'No description provided.'}",
        reply_markup=product_admin_actions_reply_menu(product.id, product.is_active),
    )


@router.message(StateFilter("*"), F.text.func(_is_product_selection))
async def admin_product_detail_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Product #")
    if not product_id:
        await message.answer("Invalid product selection.")
        return
    rows = await list_all_products(session)
    product_row = next((row for row in rows if row[0].id == product_id), None)
    if not product_row:
        await message.answer("Product not found.", reply_markup=admin_reply_menu())
        return
    product, stock_count = product_row
    await message.answer(
        f"Product Details\n\n"
        f"Product ID: #{product.id}\n"
        f"Name: {product.name}\n"
        f"Price: {money(product.price)}\n"
        f"Available Stock: {stock_count}\n"
        f"Status: {'Active' if product.is_active else 'Disabled'}\n\n"
        f"Description: {product.description or 'No description provided.'}",
        reply_markup=product_admin_actions_reply_menu(product.id, product.is_active),
    )


@router.callback_query(F.data.startswith("admin_product:"))
async def admin_product_detail(callback: CallbackQuery, session: AsyncSession) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    product_id = int(callback.data.split(":", 1)[1])
    rows = await list_all_products(session)
    product_row = next((row for row in rows if row[0].id == product_id), None)
    if not product_row:
        await callback.answer("Product not found.", show_alert=True)
        return
    product, stock_count = product_row
    await callback.message.edit_text(
        f"Product #{product.id}\n\n"
        f"Name: {product.name}\n"
        f"Price: {money(product.price)}\n"
        f"Stock: {stock_count}\n"
        f"Status: {'active' if product.is_active else 'disabled'}\n\n"
        f"{product.description}",
    )
    await callback.message.answer(
        "Choose product action.",
        reply_markup=product_admin_actions_reply_menu(product.id, product.is_active),
    )
    await callback.answer()


@router.message(StateFilter("*"), F.text.in_({"Add Product", "ADD PRODUCT", "➕ Add Product"}))
async def add_product_start_text(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await state.set_state(ProductForm.details)
    await message.answer(
        "Add Product\n\nSend product details in this format:\n\n"
        "Name | Price | Description\n\n"
        "Example:\nGmail Fresh | 120 | Fresh Gmail accounts"
    )


@router.callback_query(F.data == "admin_add_product")
async def add_product_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    await state.set_state(ProductForm.details)
    await callback.message.edit_text(
        "Add Product\n\nSend product details in this format:\n\n"
        "Name | Price | Description\n\n"
        "Example:\nGmail Fresh | 2.50 | Fresh Gmail accounts"
    )
    await callback.answer()


@router.message(StateFilter("*"), F.text.in_({"Add Stock", "ADD STOCK", "📥 Add Stock"}))
async def add_stock_start_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    products = await list_all_products(session)
    if not products:
        await message.answer("Create a product before adding stock.", reply_markup=admin_reply_menu())
    else:
        await state.set_state(StockForm.product_id)
        await message.answer(
            "Send the product ID to stock.\n\n"
            + "\n".join(f"#{product.id} {product.name}" for product, _ in products)
        )


@router.message(ProductForm.details)
async def add_product_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    parts = [part.strip() for part in message.text.split("|", 2)]
    if len(parts) != 3:
        await message.answer("Invalid format.\n\nUse: Name | Price | Description")
        return
    try:
        product = await create_product(session, parts[0], float(parts[1]), parts[2])
    except ValueError:
        await message.answer("Product price must be a valid number.")
        return
    await state.clear()
    await message.answer(
        f"Product Created\n\nProduct ID: #{product.id}\nName: {product.name}\nPrice: {money(product.price)}",
        reply_markup=product_admin_actions_reply_menu(product.id, True),
    )


@router.message(StateFilter("*"), F.text.func(_is_edit_product_action))
async def edit_product_start_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Edit Product #")
    if not product_id:
        await message.answer("Invalid product action.")
        return
    rows = await list_all_products(session)
    product_row = next((row for row in rows if row[0].id == product_id), None)
    if not product_row:
        await message.answer("Product not found.", reply_markup=admin_reply_menu())
        return
    product, _ = product_row
    await state.update_data(product_id=product.id)
    await state.set_state(ProductEditForm.details)
    await message.answer(
        "Edit Product\n\n"
        "Send updated product details in this format:\n\n"
        "Name | Price | Description\n\n"
        "Current:\n"
        f"{product.name} | {float(product.price):.2f} | {product.description or ''}"
    )


@router.message(ProductEditForm.details)
async def edit_product_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    parts = [part.strip() for part in message.text.split("|", 2)]
    if len(parts) != 3:
        await message.answer("Invalid format.\n\nUse: Name | Price | Description")
        return
    try:
        price = float(parts[1])
    except ValueError:
        await message.answer("Product price must be a valid number.")
        return

    data = await state.get_data()
    try:
        product = await update_product(session, int(data["product_id"]), parts[0], price, parts[2])
    except ValueError as exc:
        await message.answer(str(exc))
        return
    await state.clear()
    if not product:
        await message.answer("Product not found.", reply_markup=admin_reply_menu())
        return
    await message.answer(
        f"Product Updated\n\nProduct ID: #{product.id}\nName: {product.name}\nPrice: {money(product.price)}",
        reply_markup=product_admin_actions_reply_menu(product.id, product.is_active),
    )


@router.callback_query(F.data == "admin_add_stock")
async def add_stock_start(callback: CallbackQuery, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    products = await list_all_products(session)
    if not products:
        await callback.message.edit_text("Create a product before adding stock.")
        await callback.message.answer("Admin Panel", reply_markup=admin_reply_menu())
    else:
        await state.set_state(StockForm.product_id)
        await callback.message.edit_text(
            "Send the product ID to stock.\n\n"
            + "\n".join(f"#{product.id} {product.name}" for product, _ in products)
        )
    await callback.answer()


@router.message(StateFilter("*"), F.text.func(_is_add_stock_action))
async def add_stock_for_product_text(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Add Stock #")
    if not product_id:
        await message.answer("Invalid product action.")
        return
    await state.update_data(product_id=product_id)
    await state.set_state(StockForm.payload)
    await message.answer(
        "Send bulk stock lines, one item per line, or upload .xlsx/.csv/.txt.\n\n"
        "email1@example.com|password1\nemail2@example.com|password2"
    )


@router.message(StateFilter("*"), F.text.func(_is_export_stock_action))
async def export_stock_text(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Export Stock #")
    if not product_id:
        await message.answer("Invalid product action.", reply_markup=admin_reply_menu())
        return
    rows = [["Stock ID", "Product ID", "Payload", "Created"]]
    for item in await unsold_stock_items(session, product_id):
        rows.append([item.id, item.product_id, item.payload, str(item.created_at)])
    await log_admin_action(session, message.from_user.id, "export_stock", "product", product_id, details=f"items={len(rows) - 1}")
    await message.answer_document(
        _xlsx_file(f"product_{product_id}_unsold_stock.xlsx", rows, "Unsold Stock"),
        caption=f"Unsold stock export ready for product #{product_id}.",
        reply_markup=admin_reply_menu(),
    )


@router.message(StateFilter("*"), F.text.func(_is_import_stock_url_action))
async def import_stock_url_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Import Stock URL #")
    if not product_id:
        await message.answer("Invalid product action.", reply_markup=admin_reply_menu())
        return
    await state.update_data(product_id=product_id)
    await state.set_state(StockUrlForm.url)
    await message.answer(
        f"Import Stock URL\n\n"
        f"Product ID: #{product_id}\n\n"
        "Send a direct .txt or .csv download URL.\n\n"
        "Supported format:\n"
        "email1@example.com|password1\n"
        "email2@example.com|password2\n\n"
        "Dropbox tip: share link should end with dl=1."
    )


@router.message(StateFilter("*"), F.text.in_({"Deposits", "DEPOSITS", "💳 Deposits"}))
async def deposits_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    rows = await pending_deposits(session)
    if not rows:
        await message.answer("Deposits\n\nNo pending deposit requests.", reply_markup=admin_reply_menu())
    else:
        first = rows[0]
        text = await _deposit_details_text(session, first, len(rows))
        if getattr(first, "proof_file_id", None):
            await message.answer_photo(first.proof_file_id, caption=f"Payment screenshot for deposit #{first.id}")
        await message.answer(text, reply_markup=deposit_review_reply_menu(first.id))


@router.message(StateFilter("*"), F.text.in_({"Replacements", "🔁 Replacements"}))
async def replacements_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    rows = await pending_replacements(session)
    if not rows:
        await message.answer("Replacements\n\nNo pending replacement requests.", reply_markup=admin_reply_menu())
        return
    first = rows[0]
    text = await _replacement_details_text(session, first, len(rows))
    if getattr(first, "proof_file_id", None) and getattr(first, "proof_file_name", None):
        await message.answer_document(first.proof_file_id, caption=f"Replacement proof for request #{first.id}")
    elif getattr(first, "proof_file_id", None):
        await message.answer_photo(first.proof_file_id, caption=f"Replacement proof for request #{first.id}")
    elif getattr(first, "proof_text", None):
        await message.answer(f"Proof text #{first.id}:\n\n{first.proof_text[:3500]}")
    await message.answer(text, reply_markup=replacement_review_reply_menu(first.id))


@router.message(StateFilter("*"), F.text.func(_is_replacement_review))
async def replacement_review_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    approve = _is_replacement_approve(message.text)
    request_id = _id_from_hash_button(message.text, "Approve Replace #" if approve else "Reject Replace #")
    if not request_id:
        await message.answer("Invalid replacement action.", reply_markup=admin_reply_menu())
        return
    ok, text, request = await review_replacement(session, request_id, approve=approve)
    if not request:
        await message.answer(text, reply_markup=admin_reply_menu())
        return
    await log_admin_action(
        session,
        message.from_user.id,
        "approve_replace" if approve else "reject_replace",
        "replacement",
        request.id,
        details=f"quantity={request.quantity}; refund={float(request.refund_amount)}",
    )
    from bot.database.models import User

    user = await session.get(User, request.user_id)
    if user:
        try:
            await message.bot.send_message(
                user.telegram_id,
                ("Replace Approved\n\n" if approve else "Replace Rejected\n\n")
                + f"Request ID: #{request.id}\n"
                + f"Quantity: {request.quantity}\n"
                + f"Refund Added: {money(request.refund_amount)}\n\n"
                + ("Your balance has been updated." if approve else "Admin checked your proof and rejected this request."),
            )
        except Exception:
            pass
    await _send_next_replacement_after_review(message, session, request)


def _auto_stock_status_text(source: object | None, current_stock: int, product_id: int) -> str:
    if not source:
        return (
            "Auto Stock Status\n\n"
            f"Product ID: #{product_id}\n"
            f"Current Stock: {current_stock}\n"
            "Status: Not configured\n\n"
            "Use Auto Refill and send:\n"
            "DropboxLink | 40000 | 20000"
        )
    status = "Active" if source.is_active else "Stopped"
    last_run = source.last_run_at or "Never"
    last_error = source.last_error or "None"
    return (
        "Auto Stock Status\n\n"
        f"Product ID: #{product_id}\n"
        f"Current Stock: {current_stock}\n"
        f"Status: {status}\n"
        f"Target Stock: {source.target_stock}\n"
        f"Refill When Stock <= {source.refill_threshold}\n"
        f"Next Source Line: {source.next_line_number}\n"
        f"Last Added: {source.last_added_count}\n"
        f"Last Run: {last_run}\n"
        f"Last Error: {last_error}"
    )


@router.message(StateFilter("*"), F.text.func(_is_auto_refill_action))
async def auto_refill_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Auto Refill #")
    if not product_id:
        await message.answer("Invalid product action.", reply_markup=admin_reply_menu())
        return
    await state.update_data(product_id=product_id)
    await state.set_state(AutoStockForm.details)
    await message.answer(
        "Auto Stock Refill\n\n"
        f"Product ID: #{product_id}\n\n"
        "Send Dropbox/direct URL like this:\n\n"
        "DropboxLink | 40000 | 20000\n\n"
        "40000 = keep stock up to this number\n"
        "20000 = refill when stock becomes this number or lower\n\n"
        "You can also send only the link. Default: 40000 target, 20000 threshold."
    )


@router.message(StateFilter("*"), F.text.func(_is_auto_status_action))
async def auto_refill_status(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Auto Status #")
    if not product_id:
        await message.answer("Invalid product action.", reply_markup=admin_reply_menu())
        return
    source = await get_auto_stock_source(session, product_id)
    current_stock = await unsold_stock_count(session, product_id)
    await message.answer(
        _auto_stock_status_text(source, current_stock, product_id),
        reply_markup=product_admin_actions_reply_menu(product_id, True),
    )


@router.message(StateFilter("*"), F.text.func(_is_stop_auto_action))
async def auto_refill_stop(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Stop Auto #")
    if not product_id:
        await message.answer("Invalid product action.", reply_markup=admin_reply_menu())
        return
    stopped = await stop_auto_stock_source(session, product_id)
    await message.answer(
        "Auto stock refill stopped." if stopped else "Auto stock refill was not configured for this product.",
        reply_markup=product_admin_actions_reply_menu(product_id, True),
    )


@router.message(StateFilter("*"), F.text.func(_is_reset_auto_action))
async def auto_refill_reset(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Reset Auto #")
    if not product_id:
        await message.answer("Invalid product action.", reply_markup=admin_reply_menu())
        return
    reset = await reset_auto_stock_progress(session, product_id)
    await message.answer(
        "Auto stock progress reset. Next refill will start from line 1 again." if reset else "Auto stock refill was not configured for this product.",
        reply_markup=product_admin_actions_reply_menu(product_id, True),
    )


@router.callback_query(F.data.startswith("admin_stock_for:"))
async def add_stock_for_product(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    product_id = int(callback.data.split(":", 1)[1])
    await state.update_data(product_id=product_id)
    await state.set_state(StockForm.payload)
    await callback.message.edit_text(
        "Send bulk stock lines, one item per line, or upload .xlsx/.csv/.txt.\n\n"
        "email1@example.com|password1\nemail2@example.com|password2"
    )
    await callback.answer()


@router.message(StateFilter("*"), F.text.in_({"Coupons", "COUPONS", "🏷 Coupons"}))
async def add_coupon_start_text(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    await state.set_state(CouponAdminForm.details)
    await message.answer(
        "Send coupon details:\n\n"
        "CODE | Amount | Max Uses\n\n"
        "Example:\nWELCOME10 | 100 | 100"
    )


@router.message(StockForm.product_id)
async def stock_product_id(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    try:
        product_id = int(message.text)
    except (TypeError, ValueError):
        await message.answer("Send a numeric product ID.")
        return
    await state.update_data(product_id=product_id)
    await state.set_state(StockForm.payload)
    await message.answer(
        "Send bulk stock lines, one item per line, or upload .xlsx/.csv/.txt.\n\n"
        "email1@example.com|password1\nemail2@example.com|password2"
    )


async def _process_stock_document(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return

    document = message.document
    if not document:
        await message.answer("Please upload a .xlsx, .csv, or .txt stock file.")
        return
    file_name = document.file_name or ""
    if not file_name.lower().endswith(SUPPORTED_STOCK_EXTENSIONS):
        await message.answer("Unsupported file type. Please upload only .xlsx, .csv, or .txt stock files.")
        return
    if document.file_size and document.file_size > MAX_TELEGRAM_DOWNLOAD_BYTES:
        await message.answer(
            "Stock file is too large for Telegram bot download.\n\n"
            "Please split the stock into smaller files and upload again.\n\n"
            "Best format:\n"
            "email1@example.com|password1\n"
            "email2@example.com|password2\n\n"
            "Tip: .txt or .csv files are lighter than .xlsx.",
            reply_markup=admin_reply_menu(),
        )
        return

    data = await state.get_data()
    product_id = data.get("product_id")
    if not product_id:
        await state.clear()
        await message.answer("Product ID missing. Please open Products, select a product, then tap Add Stock again.", reply_markup=admin_reply_menu())
        return

    await message.answer(f"Processing stock file...\n\nFile: {file_name}")
    buffer = BytesIO()
    try:
        await message.bot.download(document, destination=buffer)
        lines = parse_stock_file(file_name, buffer.getvalue())
    except Exception as exc:
        reason = str(exc)
        if "file is too big" in reason.lower():
            await message.answer(
                "Stock file is too large for Telegram bot download.\n\n"
                "Please split the stock into smaller files, or upload as a lighter .txt/.csv file.\n\n"
                "Format:\n"
                "email1@example.com|password1\n"
                "email2@example.com|password2",
                reply_markup=admin_reply_menu(),
            )
            return
        await message.answer(f"Could not read this stock file.\n\nReason: {reason}")
        return

    if not lines:
        await message.answer("No valid stock lines found in this file.")
        return

    try:
        count = await add_stock(session, int(product_id), lines)
    except Exception as exc:
        await message.answer(f"Stock upload failed while saving to database.\n\nReason: {exc}")
        return
    await state.clear()
    await message.answer(
        f"Stock Uploaded\n\nFile: {file_name}\nAdded Items: {count}",
        reply_markup=admin_reply_menu(),
    )


@router.message(StockForm.payload, F.document)
async def stock_payload_file(message: Message, state: FSMContext, session: AsyncSession) -> None:
    await _process_stock_document(message, state, session)


@router.message(StockForm.payload)
async def stock_payload(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    if not message.text:
        await message.answer("Send stock text or upload a .xlsx/.csv/.txt file.")
        return
    data = await state.get_data()
    product_id = data.get("product_id")
    if not product_id:
        await state.clear()
        await message.answer("Product ID missing. Please open Products, select a product, then tap Add Stock again.", reply_markup=admin_reply_menu())
        return
    count = await add_stock(session, int(product_id), message.text.splitlines())
    await state.clear()
    await message.answer(f"Stock Added\n\nAdded Items: {count}", reply_markup=admin_reply_menu())


@router.message(AutoStockForm.details)
async def auto_refill_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    if not message.text:
        await message.answer("Please send the Dropbox/direct URL.")
        return

    data = await state.get_data()
    product_id = int(data["product_id"])
    parts = [part.strip() for part in message.text.split("|")]
    url = parts[0]
    if not url.startswith(("http://", "https://")):
        await message.answer("Please send a valid http/https Dropbox or direct download URL.")
        return
    try:
        target_stock = int(parts[1]) if len(parts) > 1 and parts[1] else DEFAULT_TARGET_STOCK
        refill_threshold = int(parts[2]) if len(parts) > 2 and parts[2] else DEFAULT_REFILL_THRESHOLD
    except ValueError:
        await message.answer("Target and threshold must be whole numbers.\n\nExample:\nDropboxLink | 40000 | 20000")
        return
    if target_stock < 1000:
        await message.answer("Target stock should be at least 1000.")
        return
    if refill_threshold >= target_stock:
        await message.answer("Threshold must be smaller than target stock.\n\nExample:\nDropboxLink | 40000 | 20000")
        return

    source = await upsert_auto_stock_source(session, product_id, url, target_stock, refill_threshold)
    await state.clear()
    await log_admin_action(
        session,
        message.from_user.id,
        "auto_stock_setup",
        "product",
        product_id,
        details=f"target={target_stock}, threshold={refill_threshold}",
    )
    await message.answer(
        "Auto refill saved.\n\n"
        "Trying the first refill now. Large Dropbox files can take a few minutes."
    )
    try:
        result = await refill_source(session, source.id)
    except Exception as exc:
        await message.answer(
            f"Auto refill saved, but first refill failed.\n\nReason: {exc}",
            reply_markup=product_admin_actions_reply_menu(product_id, True),
        )
        return
    await message.answer(
        "Auto Refill Ready\n\n"
        f"Product ID: #{product_id}\n"
        f"Added Now: {result.added_count}\n"
        f"Current Stock: {result.current_stock}\n"
        f"Target Stock: {result.target_stock}\n"
        f"Refill When Stock <= {result.refill_threshold}\n"
        f"Next Source Line: {result.next_line_number}\n\n"
        "Bot will check automatically every few minutes.",
        reply_markup=product_admin_actions_reply_menu(product_id, True),
    )


@router.message(StockUrlForm.url)
async def stock_url_import_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    url = (message.text or "").strip()
    if not url.startswith(("http://", "https://")):
        await message.answer("Please send a valid http/https download URL.")
        return
    data = await state.get_data()
    product_id = int(data["product_id"])
    await message.answer(
        "Stock URL import started.\n\n"
        "Please wait. Large files can take a while. Do not send another import for this product until it finishes."
    )
    try:
        count = await import_stock_from_url(session, product_id, url)
    except Exception as exc:
        await message.answer(f"Stock URL import failed.\n\nReason: {exc}", reply_markup=product_admin_actions_reply_menu(product_id, True))
        return
    await state.clear()
    await log_admin_action(session, message.from_user.id, "import_stock_url", "product", product_id, details=f"items={count}")
    await message.answer(
        f"Stock URL Import Completed\n\nProduct ID: #{product_id}\nAdded Items: {count}",
        reply_markup=product_admin_actions_reply_menu(product_id, True),
    )


@router.message(StateFilter("*"), F.document)
async def admin_document_without_stock_state(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    if data.get("product_id"):
        await _process_stock_document(message, state, session)
        return
    await message.answer(
        "File received, but no stock upload session is active.\n\n"
        "Go to Products > select product > Add Stock, then upload the file again.",
        reply_markup=admin_reply_menu(),
    )


@router.callback_query(F.data.startswith("admin_toggle_product:"))
async def admin_toggle(callback: CallbackQuery, session: AsyncSession) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    product_id = int(callback.data.split(":", 1)[1])
    product = await toggle_product(session, product_id)
    if not product:
        await callback.answer("Product not found.", show_alert=True)
        return
    await callback.message.edit_text(
        f"{product.name} is now {'active' if product.is_active else 'disabled'}.",
    )
    await callback.message.answer(
        "Choose product action.",
        reply_markup=product_admin_actions_reply_menu(product.id, product.is_active),
    )
    await callback.answer()


@router.message(
    StateFilter("*"),
    F.text.func(_is_product_action),
)
async def toggle_or_delete_product_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return

    if _starts_with_any(message.text, ("Enable Product #", "Disable Product #")):
        product_id = _id_from_hash_button(message.text, "Enable Product #") or _id_from_hash_button(
            message.text, "Disable Product #"
        )
        if not product_id:
            await message.answer("Invalid product action.")
            return
        product = await toggle_product(session, product_id)
        if not product:
            await message.answer("Product not found.", reply_markup=admin_reply_menu())
            return
        await message.answer(
            f"{product.name} is now {'active' if product.is_active else 'disabled'}.",
            reply_markup=product_admin_actions_reply_menu(product.id, product.is_active),
        )
        return

    if _starts_with_any(message.text, ("Delete Product #",)):
        product_id = _id_from_hash_button(message.text, "Delete Product #")
        if not product_id:
            await message.answer("Invalid product action.")
            return
        await message.answer(
            "Delete this product?\n\n"
            "If this product already has orders, it will be archived from the product list and old order history will stay safe.",
            reply_markup=delete_product_confirm_reply_menu(product_id),
        )
        return

    if _starts_with_any(message.text, ("Cancel Product #",)):
        await message.answer("Cancelled.", reply_markup=admin_reply_menu())


@router.callback_query(F.data.startswith("admin_delete_product:"))
async def admin_delete_product(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    product_id = int(callback.data.split(":", 1)[1])
    await callback.message.edit_text(
        "Delete this product?\n\n"
        "If this product already has orders, it will be archived from the product list and old order history will stay safe.",
    )
    await callback.message.answer("Confirm delete.", reply_markup=delete_product_confirm_reply_menu(product_id))
    await callback.answer()


@router.callback_query(F.data.startswith("admin_delete_product_confirm:"))
async def admin_delete_product_finish(callback: CallbackQuery, session: AsyncSession) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    product_id = int(callback.data.split(":", 1)[1])
    ok, text = await delete_product(session, product_id)
    if not ok:
        await callback.answer(text, show_alert=True)
        return
    await log_admin_action(session, callback.from_user.id, "delete_product", "product", product_id, details=text)
    await callback.message.edit_text(text)
    await callback.message.answer("Admin Panel", reply_markup=admin_reply_menu())
    await callback.answer()


@router.message(
    StateFilter("*"),
    F.text.func(_is_delete_confirm),
)
async def admin_delete_product_finish_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    product_id = _id_from_hash_button(message.text, "Confirm Delete Product #")
    if not product_id:
        await message.answer("Invalid product action.")
        return
    ok, text = await delete_product(session, product_id)
    await log_admin_action(session, message.from_user.id, "delete_product", "product", product_id, details=text)
    await message.answer(text, reply_markup=admin_reply_menu())


@router.callback_query(F.data == "admin_deposits")
async def deposits(callback: CallbackQuery, session: AsyncSession) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    rows = await pending_deposits(session)
    if not rows:
        await callback.message.edit_text("Deposits\n\nNo pending deposit requests.")
        await callback.message.answer("Admin Panel", reply_markup=admin_reply_menu())
    else:
        first = rows[0]
        text = await _deposit_details_text(session, first, len(rows))
        await callback.message.edit_text("Review deposit.")
        if getattr(first, "proof_file_id", None):
            await callback.message.answer_photo(first.proof_file_id, caption=f"Payment screenshot for deposit #{first.id}")
        await callback.message.answer(text, reply_markup=deposit_review_reply_menu(first.id))
    await callback.answer()


@router.callback_query(F.data.startswith("deposit_approve:") | F.data.startswith("deposit_reject:"))
async def review_deposit_callback(callback: CallbackQuery, session: AsyncSession, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    action, raw_id = callback.data.split(":", 1)
    if action == "deposit_reject":
        await state.update_data(deposit_id=int(raw_id))
        await state.set_state(DepositRejectReasonForm.reason)
        await callback.message.answer(f"Reject Reason\n\nSend reason for deposit #{raw_id} rejection.")
        await callback.answer()
        return
    deposit = await review_deposit(session, int(raw_id), approve=action == "deposit_approve")
    if not deposit:
        await callback.answer("Deposit not found or already reviewed.", show_alert=True)
        return
    from bot.database.models import User

    user = await session.get(User, deposit.user_id)
    if user:
        approve = action == "deposit_approve"
        status_text = "approved" if approve else "rejected"
        try:
            await callback.bot.send_message(
                user.telegram_id,
                "Deposit Update\n\n"
                f"Request ID: #{deposit.id}\n"
                f"Amount: {money(deposit.amount)}\n"
                f"Status: {status_text}\n\n"
                + ("Your balance has been updated." if approve else "Please contact support if this was a mistake."),
            )
        except Exception:
            pass
    await log_admin_action(
        session,
        callback.from_user.id,
        "approve_deposit" if action == "deposit_approve" else "reject_deposit",
        "deposit",
        deposit.id,
        details=f"amount={float(deposit.amount)}",
    )
    await callback.message.edit_text(f"Deposit #{deposit.id} {deposit.status.value}.")
    await _send_next_deposit_after_review(callback.message, session, deposit)
    await callback.answer()


@router.message(
    StateFilter("*"),
    F.text.func(_is_deposit_review),
)
async def review_deposit_text(message: Message, session: AsyncSession, state: FSMContext) -> None:
    await state.clear()
    if not is_admin(message.from_user.id):
        await message.answer("You are not authorized.")
        return
    approve = _is_deposit_approve(message.text)
    deposit_id = _id_from_hash_button(message.text, "Approve Deposit #" if approve else "Reject Deposit #")
    if not deposit_id:
        await message.answer("Invalid deposit action.")
        return
    if not approve:
        await state.update_data(deposit_id=deposit_id)
        await state.set_state(DepositRejectReasonForm.reason)
        await message.answer(f"Reject Reason\n\nSend reason for deposit #{deposit_id} rejection.")
        return
    deposit = await review_deposit(session, deposit_id, approve=approve)
    if not deposit:
        await message.answer("Deposit not found or already reviewed.", reply_markup=admin_reply_menu())
        return
    from bot.database.models import User

    user = await session.get(User, deposit.user_id)
    if user:
        status_text = "approved" if approve else "rejected"
        try:
            await message.bot.send_message(
                user.telegram_id,
                "Deposit Update\n\n"
                f"Request ID: #{deposit.id}\n"
                f"Amount: {money(deposit.amount)}\n"
                f"Status: {status_text}\n\n"
                + ("Your balance has been updated." if approve else "Please contact support if this was a mistake."),
            )
        except Exception:
            pass
    await log_admin_action(
        session,
        message.from_user.id,
        "approve_deposit" if approve else "reject_deposit",
        "deposit",
        deposit.id,
        details=f"amount={float(deposit.amount)}",
    )
    await _send_next_deposit_after_review(message, session, deposit)


@router.message(DepositRejectReasonForm.reason)
async def deposit_reject_reason_finish(message: Message, session: AsyncSession, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    reason = message.text.strip()
    if not reason:
        await message.answer("Please send a reject reason.")
        return
    data = await state.get_data()
    deposit_id = int(data["deposit_id"])
    deposit = await review_deposit(session, deposit_id, approve=False)
    await state.clear()
    if not deposit:
        await message.answer("Deposit not found or already reviewed.", reply_markup=admin_reply_menu())
        return
    from bot.database.models import User

    user = await session.get(User, deposit.user_id)
    if user:
        try:
            await message.bot.send_message(
                user.telegram_id,
                "Deposit Update\n\n"
                f"Request ID: #{deposit.id}\n"
                f"Amount: {money(deposit.amount)}\n"
                "Status: rejected\n"
                f"Reason: {reason}\n\n"
                "Please contact support if this was a mistake.",
            )
        except Exception:
            pass
    await log_admin_action(
        session,
        message.from_user.id,
        "reject_deposit",
        "deposit",
        deposit.id,
        details=f"amount={float(deposit.amount)}; reason={reason}",
    )
    await _send_next_deposit_after_review(message, session, deposit)


@router.callback_query(F.data == "admin_add_coupon")
async def add_coupon_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Unauthorized.", show_alert=True)
        return
    await state.set_state(CouponAdminForm.details)
    await callback.message.edit_text(
        "Send coupon details:\n\n"
        "CODE | Amount | Max Uses\n\n"
        "Example:\nWELCOME10 | 1.00 | 100"
    )
    await callback.answer()


@router.message(CouponAdminForm.details)
async def add_coupon_finish(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if not is_admin(message.from_user.id):
        return
    parts = [part.strip() for part in message.text.split("|")]
    if len(parts) != 3:
        await message.answer("Invalid format. Use: CODE | Amount | Max Uses")
        return
    try:
        coupon = await create_coupon(session, parts[0], float(parts[1]), int(parts[2]))
    except ValueError:
        await message.answer("Amount and max uses must be numeric.")
        return
    await state.clear()
    await message.answer(
        f"Created coupon {coupon.code} worth {money(coupon.amount)}.",
        reply_markup=admin_reply_menu(),
    )
